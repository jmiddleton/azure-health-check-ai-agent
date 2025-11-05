from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import logging
import os

load_dotenv()

logger = logging.getLogger(__name__)
logger.setLevel(os.getenv("LOG_LEVEL", "INFO"))

class UpdateHealthCheckExcel:
    def __init__(self, excel_file: str):
        self.excel_file = Path(excel_file)

    def _set_color(self, cell, status: str) -> None:
        """Set the font color of a cell based on status."""
        status = status.lower()
        if status == "succeeded" or status == "no runs":
            cell.font = Font(color="00b050")  # Green
        elif "failed" in status:
            cell.font = Font(color="ed7d31")  # Orange
        elif "invalid" in status:
            cell.font = Font(color="ed7d31")  # Red
        else:
            cell.font = Font(color="000000")  # Black

    def update_sheet(self, results: list[dict]) -> None:
        current_month = datetime.now().strftime("%b %Y")
        today_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_str = today_dt.strftime("%Y-%m-%d 00:00:00")

        # --- Load workbook and sheet ---
        wb = load_workbook(self.excel_file)
        if current_month not in wb.sheetnames:
            raise ValueError(f"Sheet '{current_month}' not found in workbook.")
        ws = wb[current_month]

        # --- Identify headers (without writing to Excel) ---
        headers = [cell.value for cell in ws[1]]

        # Find System column
        try:
            col_system = next(i + 1 for i, v in enumerate(headers) if str(v).strip().lower() == "system")
        except StopIteration:
            raise ValueError("Missing 'System' column in Excel")

        # Find today's date column (match either datetime or formatted string)
        col_today = None
        for i, v in enumerate(headers, start=1):
            if isinstance(v, datetime) and v.date() == today_dt.date():
                col_today = i
                break
            if isinstance(v, str) and v.startswith(today_str[:10]):
                col_today = i
                break

        if not col_today:
            raise ValueError(f"Column for today's date ({today_str}) not found in sheet '{current_month}'")

        updated_rows = 0
        systems_in_sheet = {}

        # Map system names to row numbers
        for row in range(2, ws.max_row + 1):
            sys_val = ws.cell(row=row, column=col_system).value
            if sys_val:
                systems_in_sheet[str(sys_val).strip().lower()] = row

        # --- Update cells ---
        for r in results:
            p_name = r.get("system")
            status = r.get("status", "-")
            if not p_name:
                continue

            match_row = systems_in_sheet.get(p_name.strip().lower())
            if match_row:
                cell = ws.cell(row=match_row, column=col_today, value=status)
                self._set_color(cell, status)
                updated_rows += 1
                logger.info(f"Updated '{p_name}' → {status}")
            else:
                logger.warning(f"No match found in Excel for system '{p_name}'")

        wb.save(self.excel_file)
        logger.info(f"✅ Updated {updated_rows} rows in sheet '{current_month}' for column {today_str}")