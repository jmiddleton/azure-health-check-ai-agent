import requests
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import logging
import os
from azure.identity import DefaultAzureCredential

load_dotenv()

logger = logging.getLogger(__name__)
logger.setLevel(os.getenv("LOG_LEVEL", "INFO"))

class UpdateHealthCheckExcel:
    def __init__(self, credential: DefaultAzureCredential, file_path: str):
        self.credential = credential
        self.file_path = Path(file_path)
        self.token = self.credential.get_token("https://graph.microsoft.com/.default").token

    def _set_color(self, cell, status: str) -> None:
        """Set the font color of a cell based on status."""
        status = status.lower()
        if status == "succeeded" or status == "no runs":
            cell.font = Font(color="00b050")  # Green
        elif "failed" in status:
            cell.font = Font(color="ed7d31")  # Orange
        elif "invalid" in status:
            cell.font = Font(color="ed7d31")  # Red
        else:
            cell.font = Font(color="000000")  # Black

    def _download_excel(self) -> BytesIO:
        """Download the Excel file from OneDrive using Microsoft Graph API."""
        site_url = "https://puntanegra.sharepoint.com/sites/IntegrationTeam2-BAU" #TODO: externalize this

        headers = {"Authorization": f"Bearer {self.token}"}
        site_url = "https://graph.microsoft.com/v1.0/sites/puntanegra.sharepoint.com:/sites/IntegrationTeam2-BAU"
        site_resp = requests.get(site_url, headers=headers).json()
        print(site_resp)

        drive_resp = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_resp['id']}/drives",
            headers={"Authorization": f"Bearer {self.token}"}
        ).json()

        drive_id = drive_resp["value"][0]["id"]  # usually the default document library
        download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{self.file_path}:/content"

        resp = requests.get(download_url, headers={"Authorization": f"Bearer {self.token}"})
        if resp.status_code == 200:
            return BytesIO(resp.content)
        else:
            logger.error(f"Failed to download Excel file: {resp.status_code} {resp.text}")
            return None

    def update_sheet(self, results: list[dict]) -> None:
        current_month = datetime.now().strftime("%b %Y")
        today_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_str = today_dt.strftime("%Y-%m-%d 00:00:00")

        # --- Load workbook and sheet ---
        wb = load_workbook(self._download_excel())
        if current_month not in wb.sheetnames:
            raise ValueError(f"Sheet '{current_month}' not found in workbook.")
        ws = wb[current_month]

        # --- Identify headers (without writing to Excel) ---
        headers = [cell.value for cell in ws[1]]

        # Find System column
        try:
            col_system = next(i + 1 for i, v in enumerate(headers) if str(v).strip().lower() == "system")
        except StopIteration:
            raise ValueError("Missing 'System' column in Excel")

        # Find today's date column (match either datetime or formatted string)
        col_today = None
        for i, v in enumerate(headers, start=1):
            if isinstance(v, datetime) and v.date() == today_dt.date():
                col_today = i
                break
            if isinstance(v, str) and v.startswith(today_str[:10]):
                col_today = i
                break

        if not col_today:
            raise ValueError(f"Column for today's date ({today_str}) not found in sheet '{current_month}'")

        updated_rows = 0
        systems_in_sheet = {}

        # Map system names to row numbers
        for row in range(2, ws.max_row + 1):
            sys_val = ws.cell(row=row, column=col_system).value
            if sys_val:
                systems_in_sheet[str(sys_val).strip().lower()] = row

        # --- Update cells ---
        for r in results:
            p_name = r.get("system")
            status = r.get("status", "-")
            if not p_name:
                continue

            match_row = systems_in_sheet.get(p_name.strip().lower())
            if match_row:
                cell = ws.cell(row=match_row, column=col_today, value=status)
                self._set_color(cell, status)
                updated_rows += 1
                logger.info(f"Updated '{p_name}' → {status}")
            else:
                logger.warning(f"No match found in Excel for system '{p_name}'")

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        # upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_path}:/content"
        # requests.put(upload_url, headers={"Authorization": f"Bearer {token}"}, data=bio)

        logger.info(f"✅ Updated {updated_rows} rows in sheet '{current_month}' for column {today_str}")